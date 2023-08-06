import os
from decimal import *
from importlib.resources import files

import openpyxl

from .milestone import *
from .priority import *
from .sprint_schedule import *
from .story import *
from .table_defination import *

__all__ = ['read_excel_file', 'output_to_excel_file', 'output_to_csv_file',
           'process_excel_file']


def read_excel_file(path, table_defination: list[tuple], sprint_schedule: SprintScheduleStore) -> tuple:
    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    columns = []

    max_column_index = chr(65 + sheet.max_column - 1)

    start_column = 'A1'
    end_column = f'{max_column_index}1'

    for cells in sheet[start_column: end_column]:
        if len(columns) > 0:
            break
        for cell in cells:
            columns.append(cell.value)

    start_cell = 'A2'
    end_cell = f'{max_column_index}{sheet.max_row}'
    rows = sheet[start_cell: end_cell]

    stories = []

    for row in rows:
        story: Story = Story()
        for column_index in range(len(row)):
            column = table_defination[column_index]
            if column[2] is str:
                setattr(story, column[1], row[column_index].value)
            elif column[2] is bool:
                setattr(story, column[1], convert_to_bool(
                    row[column_index].value))
            elif column[2] is Priority:
                setattr(story, column[1], convert_to_priority(
                    row[column_index].value))
            elif column[2] is Milestone:
                milestone = Milestone(row[column_index].value)
                milestone.calc_priority(sprint_schedule)
                setattr(story, column[1], milestone)
        if story.entryDate is not None:
            stories.append(story)

    return (columns, stories)


def output_to_csv_file(file_name: str, stories: list[Story]):
    if os.path.exists(file_name):
        os.remove(file_name)

    file = open(file_name, mode='w')

    separator = '-' * 300
    for story in stories:
        file.write(f'{separator}\n')
        file.write(
            f'{story.epicJiraTicket}|{story.milestone}|{story.criticalDefect}|{story.regulatoryComplianceUrgency}\n')

    file.close()


def output_to_excel_file(file_name: str, columns: list[str], stories: list[Story], table_defination: list[tuple]):
    if os.path.exists(file_name):
        os.remove(file_name)

    wb = openpyxl.Workbook()

    sheet = wb.active

    for column_index in range(len(columns)):
        cell = sheet.cell(row=1, column=column_index + 1)
        cell.value = columns[column_index]

    for row_index in range(len(stories)):
        for column_index, column_name, _, _, _ in table_defination:
            cell = sheet.cell(row=row_index + 2, column=column_index)
            cell.value = str(getattr(stories[row_index], column_name))

    wb.save(file_name)


def process_excel_file(input_file: str, output_file: str, sprint_schedule_config: str = None, table_defination_config: str = None):
    sprint_schedule = SprintScheduleStore()
    if sprint_schedule_config is None:
        sprint_schedule.load(files('jira_tool.assets').joinpath(
            'sprint_schedule.json').read_text())
    else:
        sprint_schedule.load_file(sprint_schedule_config)

    table_defination = ExcelColumnStore()
    if table_defination_config is None:
        table_defination.load(files('jira_tool.assets').joinpath(
            'table_defination.json').read_text())
    else:
        table_defination.load_file(table_defination_config)

    excel_columns = table_defination.to_list()

    columns, stories = read_excel_file(
        input_file, excel_columns, sprint_schedule)

    sort_stories(stories, excel_columns)
    stories = sort_stories_by_override(stories)
    stories = sort_stories_by_deferred(stories)

    output_to_excel_file(output_file, columns, stories, excel_columns)
