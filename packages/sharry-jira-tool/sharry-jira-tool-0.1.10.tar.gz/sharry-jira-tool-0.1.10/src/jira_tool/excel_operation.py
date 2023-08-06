import logging
import os
import warnings
from csv import excel
from datetime import datetime
from decimal import *
from importlib.resources import files

import openpyxl

from .excel_defination import *
from .milestone import *
from .priority import *
from .sprint_schedule import *
from .story import *

__all__ = [
    "read_excel_file",
    "output_to_excel_file",
    "output_to_csv_file",
    "process_excel_file",
]

warnings.simplefilter(action="ignore", category=UserWarning)


def read_excel_file(
    path, excel_defination: list[tuple], sprint_schedule: SprintScheduleStore
) -> tuple:
    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    columns = []

    max_column_index = chr(65 + sheet.max_column - 1)

    start_column = "A1"
    end_column = f"{max_column_index}1"

    for cells in sheet[start_column:end_column]:
        if len(columns) > 0:
            break
        for cell in cells:
            columns.append(cell.value)

    start_cell = "A2"
    end_cell = f"{max_column_index}{sheet.max_row}"
    rows = sheet[start_cell:end_cell]

    stories = []

    for row in rows:
        if _should_skip(row):
            continue

        story: Story = Story()
        for column_index in range(len(row)):
            column = excel_defination[column_index]
            if column[2] is str:
                setattr(story, column[1], row[column_index].value)
            elif column[2] is bool:
                setattr(story, column[1], convert_to_bool(row[column_index].value))
            elif column[2] is Priority:
                setattr(story, column[1], convert_to_priority(row[column_index].value))
            elif column[2] is datetime:
                setattr(story, column[1], convert_to_datetime(row[column_index].value))
            elif column[2] is Milestone:
                milestone = Milestone(row[column_index].value)
                milestone.calc_priority(sprint_schedule)
                setattr(story, column[1], milestone)
        stories.append(story)

    return (columns, stories)


def _should_skip(row: list) -> bool:
    if len(row) == 0:
        return True
    else:
        first_cell_value = row[0].value
        if first_cell_value is None or len(str(first_cell_value)) == 0:
            return True
    return False


def output_to_csv_file(file_name: str, stories: list[Story]):
    if os.path.exists(file_name):
        os.remove(file_name)

    file = open(file_name, mode="w")

    separator = "-" * 300
    for story in stories:
        file.write(f"{separator}\n")
        file.write(
            f"{story.epicJiraTicket}|{story.milestone}|{story.criticalDefect}|{story.regulatoryComplianceUrgency}\n"
        )
    file.close()


def output_to_excel_file(
    file_name: str,
    columns_in_excel: list[str],
    stories: list[Story],
    excel_defination: list[tuple],
):
    if os.path.exists(file_name):
        os.remove(file_name)

    wb = openpyxl.Workbook()

    sheet = wb.active

    for column_index in range(len(columns_in_excel)):
        cell = sheet.cell(row=1, column=column_index + 1)
        cell.value = columns_in_excel[column_index]

    for row_index in range(len(stories)):
        for column_index, column_name, _, _, _ in excel_defination:
            cell = sheet.cell(row=row_index + 2, column=column_index)
            cell.value = stories[row_index].get_value(column_name)

    wb.save(file_name)


def process_excel_file(
    input_file: str,
    output_file: str,
    sprint_schedule_config: str = None,
    excel_defination_config: str = None,
):
    logger = logging.getLogger()

    sprint_schedule = SprintScheduleStore()
    if sprint_schedule_config is None:
        sprint_schedule.load(
            files("jira_tool.assets").joinpath("sprint_schedule.json").read_text()
        )
        logger.info("Loading default sprint schedule...")
    else:
        sprint_schedule.load_file(sprint_schedule_config)

    excel_defination = ExcelDefination()
    if excel_defination_config is None:
        excel_defination.load(
            files("jira_tool.assets").joinpath("excel_defination.json").read_text()
        )
        logger.info("Loading default excel defination...")
    else:
        excel_defination.load_file(excel_defination_config)

    excel_defination_columns = excel_defination.get_columns()

    excel_columns, stories = read_excel_file(
        input_file, excel_defination_columns, sprint_schedule
    )

    sort_stories(stories, excel_defination_columns)
    stories = sort_stories_by_override(stories)
    stories = sort_stories_by_deferred(stories)

    output_to_excel_file(output_file, excel_columns, stories, excel_defination_columns)

    logger.info("%s has been saved.", output_file)
