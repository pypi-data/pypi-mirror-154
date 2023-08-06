import os
from decimal import *
from operator import attrgetter

import openpyxl

from .milestone import *
from .priority import *
from .sprint_schedule import *
from .story import *

__all__ = ['read_excel_file', 'output_to_excel_file', 'output_to_csv_file',
           'sort_stories', 'sort_stories_by_deferred', 'sort_stories_by_override']


def read_excel_file(path, table_defination: list[tuple], sprint_schedule: SprintScheduleStore) -> tuple:
    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    columns = []

    max_column_index = chr(65 + sheet.max_column - 1)

    start_column = 'A1'
    end_column = f'{max_column_index}1'

    for cells in sheet[start_column: end_column]:
        if len(columns) > 0:
            break
        for cell in cells:
            columns.append(cell.value)

    start_cell = 'A2'
    end_cell = f'{max_column_index}{sheet.max_row}'
    rows = sheet[start_cell: end_cell]

    stories = []

    for row in rows:
        story: Story = Story()
        for column_index in range(len(row)):
            column = table_defination[column_index]
            if column[2] is str:
                setattr(story, column[1], row[column_index].value)
            elif column[2] is bool:
                setattr(story, column[1], convert_to_bool(
                    row[column_index].value))
            elif column[2] is Priority:
                setattr(story, column[1], convert_to_priority(
                    row[column_index].value))
            elif column[2] is Milestone:
                milestone = Milestone(row[column_index].value)
                milestone.calc_priority(sprint_schedule)
                setattr(story, column[1], milestone)
        if story.entryDate is not None:
            stories.append(story)

    return (columns, stories)


def sort_stories(stories: list[Story], table_defination: list[tuple]):
    sort_rule = []

    for _, column_name, _, need_sort, sort_desc_or_asc in table_defination:
        if need_sort is True:
            sort_rule.append((column_name, sort_desc_or_asc))

    _internal_sort_stories(stories, sort_rule)


def sort_stories_by_deferred(stories: list[Story]):
    pass


def sort_stories_by_override(stories: list[Story]):
    pass


def _internal_sort_stories(stories: list[Story], keys: list[tuple]):
    for key, isReversed in reversed(keys):
        stories.sort(key=attrgetter(key), reverse=isReversed)


def output_to_csv_file(file_name: str, stories: list[Story]):
    if os.path.exists(file_name):
        os.remove(file_name)

    file = open(file_name, mode='w')

    separator = '-' * 300
    for story in stories:
        file.write(f'{separator}\n')
        file.write(
            f'{story.epicJiraTicket}|{story.milestone}|{story.criticalDefect}|{story.regulatoryComplianceUrgency}\n')

    file.close()


def output_to_excel_file(file_name: str, columns: list[str], stories: list[Story], table_defination: list[tuple]):
    if os.path.exists(file_name):
        os.remove(file_name)

    wb = openpyxl.Workbook()

    sheet = wb.active

    for column_index in range(len(columns)):
        cell = sheet.cell(row=1, column=column_index + 1)
        cell.value = columns[column_index]

    for row_index in range(len(stories)):
        for column_index, column_name, _, _, _ in table_defination:
            cell = sheet.cell(row=row_index + 2, column=column_index)
            cell.value = str(getattr(stories[row_index], column_name))

    wb.save(file_name)
